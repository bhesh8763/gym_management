"""
Reports & Analytics — read-only aggregation views + export endpoints.

These endpoints don't own any models of their own; they query across the
existing apps (payments, memberships, attendance, equipment, lockers,
members, staff) and return summarized data for dashboards/charts.

All endpoints are Owner/Staff only.

Export endpoints:
    GET /api/reports/export/attendance/   ?format=csv|excel  ?start=YYYY-MM-DD &end=YYYY-MM-DD
    GET /api/reports/export/memberships/  ?format=csv|excel  ?status=ACTIVE|...
    GET /api/reports/export/revenue/      ?format=csv|excel  ?start=YYYY-MM-DD &end=YYYY-MM-DD
    GET /api/reports/export/members/      ?format=csv|excel
    GET /api/reports/export/equipment/    ?format=csv|excel
    GET /api/reports/export/maintenance/  ?format=csv|excel
    GET /api/reports/export/diet/         ?format=csv|excel  ?is_active=true
    GET /api/reports/export/progress/     ?format=csv|excel  ?member=<id>
    GET /api/reports/export/staff/        ?format=csv|excel
"""
import csv
import io
from datetime import timedelta, date

from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from apps.accounts.permissions import IsOwner
from apps.payments.models import Payment
from apps.memberships.models import Membership, MembershipPlan
from apps.attendance.models import Attendance
from apps.equipment.models import Equipment, MaintenanceRecord
from apps.lockers.models import Locker, LockerAssignment
from apps.members.models import MemberProfile
from apps.staff.models import StaffProfile, LeaveRequest
from apps.diet.models import DietPlan, Meal
from apps.progress.models import ProgressEntry, PersonalRecord


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_date(value, fallback):
    if not value:
        return fallback
    try:
        return date.fromisoformat(value)
    except ValueError:
        return fallback


def _make_csv_response(filename, headers, rows):
    """Build an HTTP response with CSV content."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def _make_excel_response(filename, headers, rows):
    """Build an HTTP response with Excel (.xlsx) content using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None  # Caller handles fallback

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.replace('.xlsx', '')

    # Header row styling
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_response(request, filename_base, headers, rows):
    """
    Return CSV or Excel depending on ?format= query param.
    Defaults to CSV if openpyxl is not installed.
    """
    fmt = request.query_params.get('format', 'csv').lower()
    if fmt == 'excel':
        response = _make_excel_response(f'{filename_base}.xlsx', headers, rows)
        if response:
            return response
        # openpyxl not installed, fall back to CSV
    return _make_csv_response(f'{filename_base}.csv', headers, rows)


# ─────────────────────────────────────────────────────────────────────────────
# Standard report views (JSON)
# ─────────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsOwner])
def revenue_report(request):
    """
    GET /api/reports/revenue/
    Monthly revenue trend (last 12 months), breakdown by payment method,
    breakdown by payment status, and breakdown by what the payment was for.
    """
    today = timezone.now().date()
    twelve_months_ago = (today.replace(day=1) - timedelta(days=365))

    monthly = (
        Payment.objects.filter(status=Payment.PaymentStatus.PAID, created_at__date__gte=twelve_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total=Sum('amount_paid'), count=Count('id'))
        .order_by('month')
    )

    by_method = (
        Payment.objects.filter(status=Payment.PaymentStatus.PAID)
        .values('payment_method')
        .annotate(total=Sum('amount_paid'), count=Count('id'))
        .order_by('-total')
    )

    by_status = (
        Payment.objects.values('status')
        .annotate(total=Sum('amount_paid'), count=Count('id'))
        .order_by('-total')
    )

    by_purpose = (
        Payment.objects.filter(status=Payment.PaymentStatus.PAID)
        .values('payment_for')
        .annotate(total=Sum('amount_paid'), count=Count('id'))
        .order_by('-total')
    )

    totals = Payment.objects.aggregate(
        collected=Sum('amount_paid', filter=Q(status=Payment.PaymentStatus.PAID)),
        pending=Sum('amount_paid', filter=Q(status=Payment.PaymentStatus.PENDING)),
        refunded=Sum('amount_paid', filter=Q(status=Payment.PaymentStatus.REFUNDED)),
    )

    return Response({
        'monthly_trend': [
            {'month': row['month'].strftime('%Y-%m'), 'total': row['total'] or 0, 'count': row['count']}
            for row in monthly
        ],
        'by_method': list(by_method),
        'by_status': list(by_status),
        'by_purpose': list(by_purpose),
        'totals': {
            'collected': totals['collected'] or 0,
            'pending': totals['pending'] or 0,
            'refunded': totals['refunded'] or 0,
        },
    })


@api_view(['GET'])
@permission_classes([IsOwner])
def membership_report(request):
    """
    GET /api/reports/memberships/
    Status breakdown, plan popularity, and memberships expiring in the
    next 30 days (useful for renewal follow-ups).
    """
    today = timezone.now().date()

    by_status = (
        Membership.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    by_plan = (
        Membership.objects.values('plan__name')
        .annotate(count=Count('id'), revenue=Sum('price_paid'))
        .order_by('-count')
    )

    expiring_soon = (
        Membership.objects.filter(
            status=Membership.Status.ACTIVE,
            end_date__gte=today,
            end_date__lte=today + timedelta(days=30),
        )
        .select_related('member', 'plan')
        .order_by('end_date')
    )

    return Response({
        'by_status': list(by_status),
        'by_plan': [
            {'plan': row['plan__name'], 'count': row['count'], 'revenue': row['revenue'] or 0}
            for row in by_plan
        ],
        'total_active': Membership.objects.filter(status=Membership.Status.ACTIVE).count(),
        'total_members': MemberProfile.objects.count(),
        'expiring_soon': [
            {
                'id': m.id,
                'member_name': m.member.get_full_name(),
                'plan_name': m.plan.name,
                'end_date': m.end_date,
            }
            for m in expiring_soon[:25]
        ],
    })


@api_view(['GET'])
@permission_classes([IsOwner])
def attendance_report(request):
    """
    GET /api/reports/attendance/
    Daily check-in counts for the last 30 days, plus a breakdown by
    attendance type (member/staff/trainer).
    """
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=29)

    daily = (
        Attendance.objects.filter(date__gte=thirty_days_ago)
        .values('date')
        .annotate(count=Count('id'))
        .order_by('date')
    )

    by_type = (
        Attendance.objects.filter(date__gte=thirty_days_ago)
        .values('attendance_type')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    return Response({
        'daily': [{'date': row['date'], 'count': row['count']} for row in daily],
        'by_type': list(by_type),
        'total_last_30_days': Attendance.objects.filter(date__gte=thirty_days_ago).count(),
    })


@api_view(['GET'])
@permission_classes([IsOwner])
def equipment_report(request):
    """
    GET /api/reports/equipment/
    Condition breakdown, maintenance cost total, and upcoming/overdue
    maintenance counts.
    """
    today = timezone.now().date()

    by_condition = (
        Equipment.objects.values('condition')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    maintenance_cost = MaintenanceRecord.objects.aggregate(total=Sum('cost'))['total'] or 0

    upcoming = MaintenanceRecord.objects.filter(
        status=MaintenanceRecord.MaintenanceStatus.SCHEDULED,
        scheduled_date__gte=today,
    ).count()

    overdue = MaintenanceRecord.objects.filter(
        status=MaintenanceRecord.MaintenanceStatus.SCHEDULED,
        scheduled_date__lt=today,
    ).count()

    return Response({
        'by_condition': list(by_condition),
        'total_items': Equipment.objects.aggregate(total=Sum('quantity'))['total'] or 0,
        'maintenance_cost_total': maintenance_cost,
        'upcoming_maintenance': upcoming,
        'overdue_maintenance': overdue,
    })


@api_view(['GET'])
@permission_classes([IsOwner])
def locker_report(request):
    """
    GET /api/reports/lockers/
    Occupancy breakdown and monthly recurring revenue from locker fees.
    """
    by_status = (
        Locker.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    total = Locker.objects.count()
    occupied = Locker.objects.filter(status=Locker.LockerStatus.OCCUPIED).count()

    mrr = (
        LockerAssignment.objects.filter(is_active=True)
        .select_related('locker')
        .aggregate(total=Sum('locker__monthly_fee'))['total'] or 0
    )

    return Response({
        'by_status': list(by_status),
        'total_lockers': total,
        'occupied': occupied,
        'occupancy_rate': round((occupied / total * 100), 1) if total else 0,
        'monthly_recurring_revenue': mrr,
    })


@api_view(['GET'])
@permission_classes([IsOwner])
def staff_report(request):
    """
    GET /api/reports/staff/
    Headcount by department and pending leave request count.
    """
    by_department = (
        StaffProfile.objects.values('department')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    return Response({
        'by_department': list(by_department),
        'total_staff': StaffProfile.objects.count(),
        'pending_leave_requests': LeaveRequest.objects.filter(
            status=LeaveRequest.LeaveStatus.PENDING
        ).count(),
    })


@api_view(['GET'])
@permission_classes([IsOwner])
def overview_report(request):
    """
    GET /api/reports/overview/
    Top-line KPIs for a landing dashboard: revenue this month, active
    members, occupancy, and equipment condition — one call per page load.
    """
    today = timezone.now().date()
    month_start = today.replace(day=1)

    revenue_this_month = Payment.objects.filter(
        status=Payment.PaymentStatus.PAID, created_at__date__gte=month_start,
    ).aggregate(total=Sum('amount_paid'))['total'] or 0

    return Response({
        'revenue_this_month': revenue_this_month,
        'active_memberships': Membership.objects.filter(status=Membership.Status.ACTIVE).count(),
        'total_members': MemberProfile.objects.count(),
        'total_staff': StaffProfile.objects.count(),
        'lockers_occupied': Locker.objects.filter(status=Locker.LockerStatus.OCCUPIED).count(),
        'lockers_total': Locker.objects.count(),
        'equipment_out_of_service': Equipment.objects.filter(
            condition=Equipment.Condition.OUT_OF_SERVICE
        ).count(),
        'attendance_today': Attendance.objects.filter(date=today).count(),
    })


# ─────────────────────────────────────────────────────────────────────────────
# Export endpoints (CSV / Excel)
# ─────────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsOwner])
def export_attendance(request):
    """
    GET /api/reports/export/attendance/?format=csv|excel&start=YYYY-MM-DD&end=YYYY-MM-DD
    Exports all attendance records in the given date range.
    Defaults to last 30 days if no range provided.
    """
    today = timezone.now().date()
    start = _parse_date(request.query_params.get('start'), today - timedelta(days=29))
    end = _parse_date(request.query_params.get('end'), today)

    records = (
        Attendance.objects
        .filter(date__gte=start, date__lte=end)
        .select_related('user', 'marked_by')
        .order_by('date', 'check_in')
    )

    headers = [
        'ID', 'Member Name', 'Display ID', 'Type', 'Date',
        'Status', 'Check In', 'Check Out', 'Duration (min)',
        'Marked By', 'Notes',
    ]
    rows = []
    for r in records:
        rows.append([
            r.id,
            r.user.get_full_name(),
            r.user.display_id or '',
            r.attendance_type,
            str(r.date),
            r.status,
            str(r.check_in) if r.check_in else '',
            str(r.check_out) if r.check_out else '',
            r.duration_minutes if r.duration_minutes is not None else '',
            r.marked_by.get_full_name() if r.marked_by else '',
            r.notes,
        ])

    return _export_response(request, f'attendance_{start}_{end}', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_memberships(request):
    """
    GET /api/reports/export/memberships/?format=csv|excel
    Exports all membership records.
    Optional ?status=ACTIVE|EXPIRED|FROZEN|CANCELLED to filter.
    """
    qs = Membership.objects.select_related('member', 'plan').order_by('-start_date')
    status_filter = request.query_params.get('status')
    if status_filter:
        qs = qs.filter(status=status_filter.upper())

    headers = [
        'ID', 'Member Name', 'Display ID', 'Email', 'Plan', 'Status',
        'Start Date', 'End Date', 'Price Paid (NPR)', 'Frozen',
        'Freeze Start', 'Freeze End',
    ]
    rows = []
    for m in qs:
        rows.append([
            m.id,
            m.member.get_full_name(),
            m.member.display_id or '',
            m.member.email,
            m.plan.name,
            m.status,
            str(m.start_date),
            str(m.end_date),
            float(m.price_paid),
            'Yes' if m.status == Membership.Status.FROZEN else 'No',
            str(m.freeze_start) if m.freeze_start else '',
            str(m.freeze_end) if m.freeze_end else '',
        ])

    return _export_response(request, 'memberships', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_revenue(request):
    """
    GET /api/reports/export/revenue/?format=csv|excel&start=YYYY-MM-DD&end=YYYY-MM-DD
    Exports all payment records in the given date range.
    Defaults to current month if no range provided.
    """
    today = timezone.now().date()
    start = _parse_date(request.query_params.get('start'), today.replace(day=1))
    end = _parse_date(request.query_params.get('end'), today)

    payments = (
        Payment.objects
        .filter(created_at__date__gte=start, created_at__date__lte=end)
        .select_related('member', 'collected_by', 'membership')
        .order_by('-created_at')
    )

    headers = [
        'Receipt No.', 'Member Name', 'Display ID', 'Payment For',
        'Amount (NPR)', 'Discount (NPR)', 'Amount Paid (NPR)',
        'Method', 'Status', 'Transaction ID',
        'Paid At', 'Collected By', 'Notes',
    ]
    rows = []
    for p in payments:
        rows.append([
            p.receipt_number,
            p.member.get_full_name(),
            p.member.display_id or '',
            p.payment_for,
            float(p.amount),
            float(p.discount),
            float(p.amount_paid),
            p.payment_method,
            p.status,
            p.transaction_id or '',
            str(p.paid_at) if p.paid_at else '',
            p.collected_by.get_full_name() if p.collected_by else '',
            p.notes,
        ])

    return _export_response(request, f'revenue_{start}_{end}', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_members(request):
    """
    GET /api/reports/export/members/?format=csv|excel
    Exports all member profiles with their active membership info.
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()

    members = (
        User.objects
        .filter(role=User.Role.MEMBER)
        .prefetch_related('memberships', 'memberships__plan')
        .order_by('first_name', 'last_name')
    )

    headers = [
        'Display ID', 'Full Name', 'Email', 'Phone',
        'Active Plan', 'Membership Status', 'Membership Expires',
        'Date Joined', 'Is Active',
    ]
    rows = []
    for member in members:
        active_membership = member.memberships.filter(status='ACTIVE').order_by('-start_date').first()
        rows.append([
            member.display_id or '',
            member.get_full_name(),
            member.email,
            member.phone or '',
            active_membership.plan.name if active_membership else 'None',
            active_membership.status if active_membership else 'None',
            str(active_membership.end_date) if active_membership else '',
            str(member.date_joined.date()),
            'Yes' if member.is_active else 'No',
        ])

    return _export_response(request, 'members', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_equipment(request):
    """
    GET /api/reports/export/equipment/?format=csv|excel
    Exports all equipment records with their latest maintenance info.
    """
    equipment = Equipment.objects.all().order_by('name')

    headers = [
        'ID', 'Name', 'Category', 'Brand', 'Model', 'Serial Number',
        'Quantity', 'Condition', 'Location', 'Purchase Date',
        'Purchase Price (NPR)', 'Notes',
    ]
    rows = []
    for e in equipment:
        rows.append([
            e.id,
            e.name,
            e.category,
            e.brand,
            e.model_number,
            e.serial_number or '',
            e.quantity,
            e.condition,
            e.location,
            str(e.purchase_date) if e.purchase_date else '',
            float(e.purchase_price) if e.purchase_price else '',
            e.notes,
        ])

    return _export_response(request, 'equipment', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_maintenance(request):
    """
    GET /api/reports/export/maintenance/?format=csv|excel
    Exports all equipment maintenance records.
    """
    records = (
        MaintenanceRecord.objects
        .select_related('equipment')
        .order_by('-scheduled_date')
    )

    headers = [
        'ID', 'Equipment', 'Type', 'Status', 'Scheduled Date',
        'Completed Date', 'Cost (NPR)', 'Notes',
    ]
    rows = []
    for r in records:
        rows.append([
            r.id,
            r.equipment.name,
            r.maintenance_type,
            r.status,
            str(r.scheduled_date) if r.scheduled_date else '',
            str(r.completed_date) if r.completed_date else '',
            float(r.cost) if r.cost else '',
            r.notes,
        ])

    return _export_response(request, 'maintenance', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_diet(request):
    """
    GET /api/reports/export/diet/?format=csv|excel
    Exports all diet plans with their meal counts.
    Optional ?is_active=true to filter.
    """
    qs = DietPlan.objects.select_related('member', 'created_by').order_by('-created_at')
    is_active = request.query_params.get('is_active')
    if is_active is not None:
        qs = qs.filter(is_active=is_active.lower() in ('true', '1', 'yes'))

    headers = [
        'ID', 'Plan Name', 'Member', 'Member Email', 'Created By',
        'Goal', 'Daily Calories', 'Protein (g)', 'Carbs (g)', 'Fats (g)',
        'Meals Count', 'Active', 'Start Date', 'End Date', 'Created At',
    ]
    rows = []
    for p in qs:
        rows.append([
            p.id,
            p.name,
            p.member.get_full_name(),
            p.member.email,
            p.created_by.get_full_name() if p.created_by else '',
            p.goal,
            p.daily_calories or '',
            p.protein_g or '',
            p.carbs_g or '',
            p.fats_g or '',
            p.meals.count(),
            'Yes' if p.is_active else 'No',
            str(p.start_date) if p.start_date else '',
            str(p.end_date) if p.end_date else '',
            str(p.created_at.date()),
        ])

    return _export_response(request, 'diet_plans', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_progress(request):
    """
    GET /api/reports/export/progress/?format=csv|excel
    Exports all progress entries.
    Optional ?member=<id> to filter by member.
    """
    qs = ProgressEntry.objects.select_related('member', 'recorded_by').order_by('-date')
    member_id = request.query_params.get('member')
    if member_id:
        qs = qs.filter(member_id=member_id)

    headers = [
        'ID', 'Member', 'Member Email', 'Date',
        'Weight (kg)', 'Height (cm)', 'BMI', 'Body Fat %',
        'Muscle Mass (kg)', 'Chest (cm)', 'Waist (cm)', 'Hips (cm)',
        'Bicep (cm)', 'Thigh (cm)', 'Recorded By', 'Notes',
    ]
    rows = []
    for e in qs:
        rows.append([
            e.id,
            e.member.get_full_name(),
            e.member.email,
            str(e.date),
            float(e.weight_kg) if e.weight_kg else '',
            float(e.height_cm) if e.height_cm else '',
            float(e.bmi) if e.bmi else '',
            float(e.body_fat_percentage) if e.body_fat_percentage else '',
            float(e.muscle_mass_kg) if e.muscle_mass_kg else '',
            float(e.chest_cm) if e.chest_cm else '',
            float(e.waist_cm) if e.waist_cm else '',
            float(e.hips_cm) if e.hips_cm else '',
            float(e.bicep_cm) if e.bicep_cm else '',
            float(e.thigh_cm) if e.thigh_cm else '',
            e.recorded_by.get_full_name() if e.recorded_by else '',
            e.notes,
        ])

    return _export_response(request, 'progress', headers, rows)


@api_view(['GET'])
@permission_classes([IsOwner])
def export_staff(request):
    """
    GET /api/reports/export/staff/?format=csv|excel
    Exports all staff profiles.
    """
    profiles = StaffProfile.objects.select_related('user').order_by('user__first_name')

    headers = [
        'ID', 'Full Name', 'Email', 'Phone', 'Department',
        'Joined Date', 'Salary (NPR)', 'Is Active',
    ]
    rows = []
    for s in profiles:
        user = s.user
        rows.append([
            s.id,
            user.get_full_name(),
            user.email,
            user.phone or '',
            s.department,
            str(s.joined_date) if s.joined_date else '',
            float(s.salary) if s.salary else '',
            'Yes' if user.is_active else 'No',
        ])

    return _export_response(request, 'staff', headers, rows)
